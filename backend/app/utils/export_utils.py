"""
Export utilities for generating CSV, Excel, and PDF files
"""
import csv
import io
from typing import List, Dict, Any, Optional
from datetime import date, datetime
from decimal import Decimal

# Optional imports for Excel and PDF support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def format_currency(amount: float) -> str:
    """Format amount as currency string."""
    return f"${amount:,.2f}"


def format_date(date_obj: date) -> str:
    """Format date as string."""
    if isinstance(date_obj, str):
        return date_obj
    return date_obj.strftime("%Y-%m-%d")


def export_to_csv(data: List[Dict[str, Any]], headers: List[str], filename: str) -> io.BytesIO:
    """Export data to CSV format."""
    output = io.BytesIO()
    writer = csv.writer(output)
    
    # Write headers
    writer.writerow(headers)
    
    # Write data rows
    for row in data:
        writer.writerow([row.get(header, '') for header in headers])
    
    output.seek(0)
    return output


def export_to_excel(data: List[Dict[str, Any]], headers: List[str], filename: str, title: Optional[str] = None) -> io.BytesIO:
    """Export data to Excel format with formatting."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    
    # Set title if provided
    row_num = 1
    if title:
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        row_num = 3
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write data rows
    for row_idx, row_data in enumerate(data, start=row_num + 1):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_num)
            value = row_data.get(header, '')
            
            # Format currency values
            if isinstance(value, (int, float, Decimal)) and header.lower() in ['debit', 'credit', 'balance', 'amount', 'total', 'revenue', 'expense', 'net']:
                cell.value = float(value)
                cell.number_format = '$#,##0.00'
            else:
                cell.value = value
            
            cell.border = border
            cell.alignment = Alignment(horizontal='left' if isinstance(value, str) else 'right', vertical='center')
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row_data in data:
            value = str(row_data.get(header, ''))
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_to_pdf(data: List[Dict[str, Any]], headers: List[str], filename: str, title: Optional[str] = None, business_name: Optional[str] = None) -> io.BytesIO:
    """Export data to PDF format."""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is not installed. Run: pip install reportlab")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Add title
    if title:
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.2*inch))
    
    # Add business name if provided
    if business_name:
        business_style = ParagraphStyle(
            'BusinessName',
            parent=styles['Normal'],
            fontSize=12,
            alignment=TA_CENTER,
            spaceAfter=10
        )
        story.append(Paragraph(business_name, business_style))
        story.append(Spacer(1, 0.1*inch))
    
    # Prepare table data
    table_data = [headers]
    for row in data:
        table_data.append([str(row.get(header, '')) for header in headers])
    
    # Create table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),  # Right align numeric columns
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return buffer

